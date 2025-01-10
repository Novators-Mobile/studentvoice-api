from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.http import HttpResponse
from rest_framework import status
from rest_framework.decorators import authentication_classes, permission_classes
from rest_framework.authentication import SessionAuthentication
from admin_api.authentication import BearerTokenAuthentication
from rest_framework.permissions import IsAuthenticated
from admin_api.models import University, Subject, Meeting, Teacher
from admin_api import serializers
from polls.models import PollResult
from polls import serializers as ps
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, NamedStyle
from openpyxl.chart import RadarChart, Reference
from datetime import datetime


@swagger_auto_schema(method='get',
                     responses={
                         200: '',
                         400: 'bad request'
                     })
@api_view(['GET'])
@authentication_classes([SessionAuthentication, BearerTokenAuthentication])
@permission_classes([IsAuthenticated])
def institute_to_subject(request, insitute_id):
    try:
        institute = University.objects.get(pk=insitute_id)
    except University.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    subjects = Subject.objects.filter(university=insitute_id).all()
    serializer = serializers.SubjectGetSerializer(subjects, many=True)
    data = list(map(lambda x: [x['name'], x['rating']], serializer.data))
    if len(data) == 0:
        return Response(status=status.HTTP_204_NO_CONTENT)

    header = NamedStyle(name="Заголовок документа", number_format="General",
                        font=Font(name='Times New Roman', bold=True, size=24),
                        alignment=Alignment(horizontal="left", vertical="bottom"))
    general = NamedStyle(name="Общий", number_format="General", font=Font(name='Times New Roman', bold=False, size=12),
                         alignment=Alignment(horizontal="left", vertical="bottom"))
    tableheader = NamedStyle(name="Заголовок таблицы", number_format="General",
                             font=Font(name='Times New Roman', bold=True, size=12),
                             alignment=Alignment(horizontal="left", vertical="bottom"))
    numberStyle = NamedStyle(name="Числа", number_format="0.0", font=Font(name='Times New Roman', bold=False, size=12),
                             alignment=Alignment(horizontal="left", vertical="bottom"))

    workbook = Workbook()
    sheet = workbook.active

    sheet.row_dimensions[1].height = 30
    sheet.column_dimensions['A'].width = 43

    sheet.cell(1, 1, 'Отчет по ' + institute.name).style = header
    sheet.cell(2, 1, 'Дата и время обращения: ' + datetime.now().strftime("%d.%m.%Y %H:%M")).style = general
    sheet.cell(4, 1, 'Дисциплина').style = tableheader
    sheet.cell(4, 2, 'Балл').style = tableheader

    lastrow = 5
    for row in data:
        sheet.cell(lastrow, 1, row[0]).style = general
        sheet.cell(lastrow, 2, row[1]).style = numberStyle
        lastrow += 1
    sheet.auto_filter.ref = 'A4:B' + str(lastrow)

    workbook.save("export.xlsx")
    with open('export.xlsx', 'rb') as f:
        file_data = f.read()
    response = HttpResponse(file_data, content_type='application/vnd.ms-excel',
                            headers={"Content-Disposition": f'attachment; filename="export.xlsx"'})

    return response


@swagger_auto_schema(method='get',
                     responses={
                         200: '',
                         400: 'bad request'
                     })
@api_view(['GET'])
@authentication_classes([SessionAuthentication, BearerTokenAuthentication])
@permission_classes([IsAuthenticated])
def institute_to_teacher(request, insitute_id):
    try:
        institute = University.objects.get(pk=insitute_id)
    except University.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    teachers = Teacher.objects.filter(university=insitute_id).all()
    serializer = serializers.TeacherGetSerializer(teachers, many=True)
    data = list(map(lambda x: [x['rating'], x['second_name'], x['first_name'], x['patronymic']], serializer.data))
    if len(data) == 0:
        return Response(status=status.HTTP_204_NO_CONTENT)

    header = NamedStyle(name="Заголовок документа", number_format="General",
                        font=Font(name='Times New Roman', bold=True, size=24),
                        alignment=Alignment(horizontal="left", vertical="bottom"))
    general = NamedStyle(name="Общий", number_format="General", font=Font(name='Times New Roman', bold=False, size=12),
                         alignment=Alignment(horizontal="left", vertical="bottom"))
    tableheader = NamedStyle(name="Заголовок таблицы", number_format="General",
                             font=Font(name='Times New Roman', bold=True, size=12),
                             alignment=Alignment(horizontal="left", vertical="bottom"))
    numberStyle = NamedStyle(name="Числа", number_format="0.0", font=Font(name='Times New Roman', bold=False, size=12),
                             alignment=Alignment(horizontal="left", vertical="bottom"))

    workbook = Workbook()
    sheet = workbook.active

    sheet.row_dimensions[1].height = 30
    sheet.column_dimensions['A'].width = 43

    sheet.cell(1, 1, 'Отчет по ' + institute.name).style = header
    sheet.cell(2, 1, 'Дата и время обращения: ' + datetime.now().strftime("%d.%m.%Y %H:%M")).style = general
    sheet.cell(4, 1, 'Преподаватель').style = tableheader
    sheet.cell(4, 2, 'Балл').style = tableheader

    lastrow = 5
    for row in data:
        sheet.cell(lastrow, 1, row[1] + " " + row[2] + " " + row[3]).style = general
        sheet.cell(lastrow, 2, row[0]).style = numberStyle
        lastrow += 1
    sheet.auto_filter.ref = 'A4:B' + str(lastrow)
    workbook.save("export.xlsx")
    with open('export.xlsx', 'rb') as f:
        file_data = f.read()
    response = HttpResponse(file_data, content_type='application/vnd.ms-excel',
                            headers={"Content-Disposition": f'attachment; filename="export.xlsx"'})

    return response


@swagger_auto_schema(method='get',
                     responses={
                         200: '',
                         400: 'bad request'
                     })
@api_view(['GET'])
@authentication_classes([SessionAuthentication, BearerTokenAuthentication])
@permission_classes([IsAuthenticated])
def subject_to_teacher(request, subject_id):
    try:
        subject = Subject.objects.get(pk=subject_id)
    except Subject.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    lecture_teachers = subject.lecture_teachers.all()
    practice_teachers = subject.practice_teachers.all()
    lecture_serializer = serializers.TeacherGetSerializer(lecture_teachers, many=True)
    practice_serializer = serializers.TeacherGetSerializer(practice_teachers, many=True)
    lecture_data = list(map(lambda x: [x['rating'], x['second_name'], x['first_name'], x['patronymic']], lecture_serializer.data))
    practice_data = list(map(lambda x: [x['rating'], x['second_name'], x['first_name'], x['patronymic']], practice_serializer.data))
    if len(lecture_data) == 0 and len(practice_data) == 0:
        return Response(status=status.HTTP_204_NO_CONTENT)

    header = NamedStyle(name="Заголовок документа", number_format="General",
                        font=Font(name='Times New Roman', bold=True, size=24),
                        alignment=Alignment(horizontal="left", vertical="bottom"))
    general = NamedStyle(name="Общий", number_format="General", font=Font(name='Times New Roman', bold=False, size=12),
                         alignment=Alignment(horizontal="left", vertical="bottom"))
    tableheader = NamedStyle(name="Заголовок таблицы", number_format="General",
                             font=Font(name='Times New Roman', bold=True, size=12),
                             alignment=Alignment(horizontal="left", vertical="bottom"))
    numberStyle = NamedStyle(name="Числа", number_format="0.0", font=Font(name='Times New Roman', bold=False, size=12),
                             alignment=Alignment(horizontal="left", vertical="bottom"))

    workbook = Workbook()
    sheet = workbook.active

    sheet.row_dimensions[1].height = 30
    sheet.column_dimensions['A'].width = 43
    sheet.column_dimensions['B'].width = 11

    sheet.cell(1, 1, 'Отчет по ' + subject.name).style = header
    sheet.cell(2, 1, 'Дата и время обращения: ' + datetime.now().strftime("%d.%m.%Y %H:%M")).style = general
    sheet.cell(4, 1, 'Преподаватель').style = tableheader
    sheet.cell(4, 2, 'Формат').style = tableheader
    sheet.cell(4, 3, 'Балл').style = tableheader

    lastrow = 5
    for row in lecture_data:
        sheet.cell(lastrow, 1, '{} {} {}'.format(row[1], row[2], row[3])).style = general
        sheet.cell(lastrow, 2, 'Лекции').style = general
        sheet.cell(lastrow, 3, row[0]).style = numberStyle
        lastrow += 1
    for row in practice_data:
        sheet.cell(lastrow, 1, '{} {} {}'.format(row[1], row[2], row[3])).style = general
        sheet.cell(lastrow, 2, 'Практики').style = general
        sheet.cell(lastrow, 3, row[0]).style = numberStyle
        lastrow += 1
    sheet.auto_filter.ref = 'A4:C' + str(lastrow)
    workbook.save("export.xlsx")
    with open('export.xlsx', 'rb') as f:
        file_data = f.read()
    response = HttpResponse(file_data, content_type='application/vnd.ms-excel',
                            headers={"Content-Disposition": f'attachment; filename="export.xlsx"'})

    return response


@swagger_auto_schema(method='get',
                     responses={
                         200: '',
                         204: 'no content',
                         400: 'bad request'
                     })
@api_view(['GET'])
@authentication_classes([SessionAuthentication, BearerTokenAuthentication])
@permission_classes([IsAuthenticated])
def subject_to_meeting(request, subject_id):
    try:
        subject = Subject.objects.get(pk=subject_id)
    except Subject.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    meetings = Meeting.objects.filter(subject=subject_id).all()
    serializer = serializers.MeetingGetSerializer(meetings, many=True)
    data = list(map(lambda x: [x['date'], x['type'], x['rating']], serializer.data))
    if len(data) == 0:
        return Response(status=status.HTTP_204_NO_CONTENT)

    header = NamedStyle(name="Заголовок документа", number_format="General",
                        font=Font(name='Times New Roman', bold=True, size=24),
                        alignment=Alignment(horizontal="left", vertical="bottom"))
    general = NamedStyle(name="Общий", number_format="General", font=Font(name='Times New Roman', bold=False, size=12),
                         alignment=Alignment(horizontal="left", vertical="bottom"))
    tableheader = NamedStyle(name="Заголовок таблицы", number_format="General",
                             font=Font(name='Times New Roman', bold=True, size=12),
                             alignment=Alignment(horizontal="left", vertical="bottom"))
    numberStyle = NamedStyle(name="Числа", number_format="0.0", font=Font(name='Times New Roman', bold=False, size=12),
                             alignment=Alignment(horizontal="left", vertical="bottom"))

    workbook = Workbook()
    sheet = workbook.active

    sheet.row_dimensions[1].height = 30
    sheet.column_dimensions['A'].width = 43
    sheet.column_dimensions['B'].width = 11

    sheet.cell(1, 1, 'Отчет по ' + subject.name).style = header
    sheet.cell(2, 1, 'Дата и время обращения: ' + datetime.now().strftime("%d.%m.%Y %H:%M")).style = general
    sheet.cell(4, 1, 'Пара').style = tableheader
    sheet.cell(4, 2, 'Формат').style = tableheader
    sheet.cell(4, 3, 'Балл').style = tableheader

    lastrow = 5
    for row in data:
        sheet.cell(lastrow, 1, subject.name + ' (' + datetime.strftime(datetime.fromisoformat(row[0]), "%d.%m.%y") + ')').style = general
        sheet.cell(lastrow, 2, row[1]).style = general
        sheet.cell(lastrow, 3, row[2]).style = numberStyle
        lastrow += 1
    sheet.auto_filter.ref = 'A4:C' + str(lastrow)
    workbook.save("export.xlsx")
    with open('export.xlsx', 'rb') as f:
        file_data = f.read()
    response = HttpResponse(file_data, content_type='application/vnd.ms-excel',
                            headers={"Content-Disposition": f'attachment; filename="export.xlsx"'})

    return response


@swagger_auto_schema(method='get',
                     responses={
                         200: '',
                         204: 'No content',
                         400: 'bad request'
                     })
@api_view(['GET'])
@authentication_classes([SessionAuthentication, BearerTokenAuthentication])
@permission_classes([IsAuthenticated])
def teacher_to_subject(request, teacher_id):
    try:
        teacher = Teacher.objects.get(pk=teacher_id)
    except Teacher.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    lectures = teacher.lecture_teachers.all()
    practices = teacher.practice_teachers.all()
    lectures_serializer = serializers.SubjectGetSerializer(lectures, many=True)
    practices_serializer = serializers.SubjectGetSerializer(practices, many=True)
    lectures_data = list(map(lambda x: [x['name'], x['rating']], lectures_serializer.data))
    practices_data = list(map(lambda x: [x['name'], x['rating']], practices_serializer.data))
    if len(lectures_data) == 0 and len(practices_data) == 0:
        return Response(status=status.HTTP_204_NO_CONTENT)

    header = NamedStyle(name="Заголовок документа", number_format="General",
                        font=Font(name='Times New Roman', bold=True, size=24),
                        alignment=Alignment(horizontal="left", vertical="bottom"))
    general = NamedStyle(name="Общий", number_format="General", font=Font(name='Times New Roman', bold=False, size=12),
                         alignment=Alignment(horizontal="left", vertical="bottom"))
    tableheader = NamedStyle(name="Заголовок таблицы", number_format="General",
                             font=Font(name='Times New Roman', bold=True, size=12),
                             alignment=Alignment(horizontal="left", vertical="bottom"))
    numberStyle = NamedStyle(name="Числа", number_format="0.0", font=Font(name='Times New Roman', bold=False, size=12),
                             alignment=Alignment(horizontal="left", vertical="bottom"))

    workbook = Workbook()
    sheet = workbook.active

    sheet.row_dimensions[1].height = 30
    sheet.column_dimensions['A'].width = 43
    sheet.column_dimensions['B'].width = 11

    sheet.cell(1, 1, 'Отчет по {} {} {}'.format(teacher.second_name, teacher.first_name, teacher.patronymic)).style = header
    sheet.cell(2, 1, 'Дата и время обращения: ' + datetime.now().strftime("%d.%m.%Y %H:%M")).style = general
    sheet.cell(4, 1, 'Дисциплина').style = tableheader
    sheet.cell(4, 2, 'Формат').style = tableheader
    sheet.cell(4, 3, 'Балл').style = tableheader

    lastrow = 5
    for row in lectures_data:
        sheet.cell(lastrow, 1, row[0]).style = general
        sheet.cell(lastrow, 2, 'Лекции').style = general
        sheet.cell(lastrow, 3, row[1]).style = numberStyle
        lastrow += 1
    for row in practices_data:
        sheet.cell(lastrow, 1, row[0]).style = general
        sheet.cell(lastrow, 2, 'Практики').style = general
        sheet.cell(lastrow, 3, row[1]).style = numberStyle
        lastrow += 1
    sheet.auto_filter.ref = 'A4:C' + str(lastrow)

    workbook.save("export.xlsx")
    with open('export.xlsx', 'rb') as f:
        file_data = f.read()
    response = HttpResponse(file_data, content_type='application/vnd.ms-excel',
                            headers={"Content-Disposition": f'attachment; filename="export.xlsx"'})

    return response


@swagger_auto_schema(method='get',
                     responses={
                         200: '',
                         400: 'bad request'
                     })
@api_view(['GET'])
@authentication_classes([SessionAuthentication, BearerTokenAuthentication])
@permission_classes([IsAuthenticated])
def teacher_to_meeting(request, teacher_id):
    try:
        teacher = Teacher.objects.get(pk=teacher_id)
    except Teacher.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)

    meetings = Meeting.objects.filter(teacher=teacher_id).all()
    serializer = serializers.MeetingGetSerializer(meetings, many=True)
    data = list(map(lambda x: [Subject.objects.get(pk=x['subject']).name, x['date'], x['type'], x['rating']], serializer.data))
    if len(data) == 0:
        return Response(status=status.HTTP_204_NO_CONTENT)

    header = NamedStyle(name="Заголовок документа", number_format="General",
                        font=Font(name='Times New Roman', bold=True, size=24),
                        alignment=Alignment(horizontal="left", vertical="bottom"))
    general = NamedStyle(name="Общий", number_format="General", font=Font(name='Times New Roman', bold=False, size=12),
                         alignment=Alignment(horizontal="left", vertical="bottom"))
    tableheader = NamedStyle(name="Заголовок таблицы", number_format="General",
                             font=Font(name='Times New Roman', bold=True, size=12),
                             alignment=Alignment(horizontal="left", vertical="bottom"))
    numberStyle = NamedStyle(name="Числа", number_format="0.0", font=Font(name='Times New Roman', bold=False, size=12),
                             alignment=Alignment(horizontal="left", vertical="bottom"))

    workbook = Workbook()
    sheet = workbook.active

    sheet.row_dimensions[1].height = 30
    sheet.column_dimensions['A'].width = 43
    sheet.column_dimensions['B'].width = 11

    sheet.cell(1, 1, 'Отчет по {} {} {}'.format(teacher.second_name, teacher.first_name, teacher.patronymic)).style = header
    sheet.cell(2, 1, 'Дата и время обращения: ' + datetime.now().strftime("%d.%m.%Y %H:%M")).style = general
    sheet.cell(4, 1, 'Пара').style = tableheader
    sheet.cell(4, 2, 'Формат').style = tableheader
    sheet.cell(4, 3, 'Балл').style = tableheader

    lastrow = 5
    for row in data:
        sheet.cell(lastrow, 1, row[0] + ' (' + datetime.strftime(datetime.fromisoformat(row[1]), "%d.%m.%y") + ')').style = general
        sheet.cell(lastrow, 1, row[2]).style = general
        sheet.cell(lastrow, 2, row[3]).style = numberStyle
        lastrow += 1
    sheet.auto_filter.ref = 'A4:C' + str(lastrow)

    workbook.save("export.xlsx")
    with open('export.xlsx', 'rb') as f:
        file_data = f.read()
    response = HttpResponse(file_data, content_type='application/vnd.ms-excel',
                            headers={"Content-Disposition": f'attachment; filename="export.xlsx"'})

    return response

@api_view(['GET'])
@authentication_classes([SessionAuthentication, BearerTokenAuthentication])
@permission_classes([IsAuthenticated])
def get_meeting(request, meeting_id):
    try:
        meeting = Meeting.objects.get(pk=meeting_id)
    except Meeting.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    
    header = NamedStyle(name="Заголовок документа", number_format="General",
                        font=Font(name='Times New Roman', bold=True, size=24),
                        alignment=Alignment(horizontal="left", vertical="bottom"))
    general = NamedStyle(name="Общий", number_format="General", font=Font(name='Times New Roman', bold=False, size=12),
                         alignment=Alignment(horizontal="left", vertical="bottom"))
    tableheader = NamedStyle(name="Заголовок таблицы", number_format="General",
                             font=Font(name='Times New Roman', bold=True, size=12),
                             alignment=Alignment(horizontal="left", vertical="bottom"))
    numberStyle = NamedStyle(name="Числа", number_format="0.0", font=Font(name='Times New Roman', bold=False, size=12),
                             alignment=Alignment(horizontal="left", vertical="bottom"))
    comment = NamedStyle(name="Комментарий", number_format="General",
                             font=Font(name='Times New Roman', bold=False, size=10),
                             alignment=Alignment(horizontal="left", vertical="bottom", wrap_text=True))
    
    workbook = Workbook()
    sheet = workbook.active

    sheet.row_dimensions[1].height = 30
    sheet.column_dimensions['A'].width = 43
    sheet.column_dimensions['B'].width = 43

    sheet.cell(1, 1, 'Отчет по паре').style = header
    sheet.cell(2, 1, 'Преподаватель: {} {} {}'.format(meeting.teacher.second_name, meeting.teacher.first_name, meeting.teacher.patronymic)).style = tableheader
    sheet.cell(3, 1, 'Дисциплина: ' + meeting.subject.name).style = tableheader
    sheet.cell(4, 1, 'Формат: ' + meeting.type).style = tableheader
    sheet.cell(5, 1, 'Дата и время пары: ' + meeting.date.strftime("%d.%m.%Y %H:%M")).style = tableheader
    sheet.cell(6, 1, 'Дата и время обращения: ' + datetime.now().strftime("%d.%m.%Y %H:%M")).style = tableheader
    sheet.cell(8, 1, 'Параметр').style = tableheader
    sheet.cell(8, 2, 'Балл').style = tableheader

    sheet.cell(9, 1, 'Взаимодействие с аудиторией').style = general
    sheet.cell(9, 2, meeting.poll.question1_avg_mark).style = numberStyle
    sheet.cell(10, 1, 'Информативность').style = general
    sheet.cell(10, 2, meeting.poll.question2_avg_mark).style = numberStyle
    sheet.cell(11, 1, 'Доступность').style = general
    sheet.cell(11, 2, meeting.poll.question3_avg_mark).style = numberStyle
    sheet.cell(12, 1, 'Интерес').style = general
    sheet.cell(12, 2, meeting.poll.question4_avg_mark).style = numberStyle
    sheet.cell(13, 1, 'Подача материала').style = general
    sheet.cell(13, 2, meeting.poll.question5_avg_mark).style = numberStyle

    sheet.cell(15, 1, 'Позитивные впечатления').style = tableheader
    sheet.cell(15, 2, 'Негативные впечатления').style = tableheader
    polls = PollResult.objects.filter(poll=meeting.poll.pk).all()
    serializer = ps.PollResultSerializer(polls, many=True)
    comments1 = list(map(lambda x: x['comment1'], serializer.data))
    comments2 = list(map(lambda x: x['comment2'], serializer.data))
    if len(comments1) != 0 or len(comments2) != 0:
        for i in range(16, 16 + max(len(comments1), len(comments2))):
            sheet.row_dimensions[i].height = 100
        lastrow = 16
        for row in comments1:
            sheet.cell(lastrow, 1, row).style = comment
            lastrow += 1
        lastrow = 16
        for row in comments2:
            sheet.cell(lastrow, 2, row).style = comment
            lastrow += 1

    workbook.save("export.xlsx")
    with open('export.xlsx', 'rb') as f:
        file_data = f.read()
    response = HttpResponse(file_data, content_type='application/vnd.ms-excel',
                            headers={"Content-Disposition": f'attachment; filename="export.xlsx"'})

    return response